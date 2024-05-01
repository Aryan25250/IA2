# This is a simple login for a Flask app
# A user can login with a username and password
# The username and password are stored in a database using SQLite
# The password is hashed using the werkzeug library
# The user is redirected to the home page if they successfully login
# The user is redirected to the login page if they fail to login
# The user is redirected to the login page if they try to access the home page without logging in

#  Path: Flask_001/app.py
from flask import Flask, render_template, request, redirect, url_for, flash, session  # for Flask
import sqlite3  # for database
import datetime
from werkzeug.security import generate_password_hash, check_password_hash  # for hashing passwords
import random
import string
import requests
import base64
import time
import pandas as pd
import openpyxl

client_id = '4567b3b6cc6344bf8a239b4ec47bfc68'
client_secret = 'dbb6313724c94c4083d6762d08d8a9d5'
credentials = base64.b64encode(f'{client_id}:{client_secret}'.encode()).decode()

db_key = "IA2"

app = Flask(__name__)  # create an instance of the Flask class called app

app.secret_key = 's3cr3t'  # set the secret key for the session (used to encrypt session data) - this should be a long random string
# the secret key should be stored in a separate file and imported, CSRF protection should also be enabled
def init_db():  # create the database if it doesn't exist
    with app.app_context():  # app.app_context() pushes an application context, which is required to access current_app
        db = sqlite3.connect(db_key)  # create a connection to the database
        cursor = db.cursor()  # create a cursor object to execute SQL statements
        cursor.execute('''CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY UNIQUE, username TEXT UNIQUE, password TEXT, image TEXT)''')  # create a table if it doesn't exist
        cursor.execute('''CREATE TABLE IF NOT EXISTS events (id INTEGER PRIMARY KEY UNIQUE, owner_id INTEGER, owner_name TEXT, owner_image TEXT, name TEXT, unix_time INTEGER, code TEXT, description TEXT, image TEXT)''')  # create a table if it doesn't exist

        db.commit()  # commit the changes

def init_spotify():
    if session and "token_expires" in session and session["token_expires"] > time.time():
        return

    headers = {'Authorization': f'Basic {credentials}'}
    data = {'grant_type': 'client_credentials'}

    # Make the token request
    response = requests.post('https://accounts.spotify.com/api/token', headers=headers, data=data, verify=False)

    # Check if the token request was successful
    if response.status_code == 200:
        token_data = response.json()
        session["access_token"] = token_data['access_token']
        session["token_expires"] = time.time() + 3600

    if not session["access_token"]:
        print(f"Token Error: {response.status_code} - {response.text}")

@app.route('/',methods=["GET","POST"])  # route() decorator binds a function to a URL
def home():  # this function is executed when the user visits the home page
    if request.method == "POST":
        session["selected_event"] = int(request.form["events"])

        return redirect(request.form["source"])


    with sqlite3.connect(db_key) as db:
        cursor = db.cursor()

        user_events = None

        if "user" in session and session["user"]["db_key"] == db_key and "user" in session:
            cursor.execute('SELECT * FROM user_events_'+str(session["user"]["id"]))
            user_events = cursor.fetchall()

        if session and user_events:
            session["user_events"] = user_events

            if not "selected_event" in session:
                session["selected_event"] = session["user_events"][0][0]

    return render_template('index.html',events=user_events)  # render the index.html template

@app.route('/games', methods=['GET', 'POST'])
def games():
    if request.method == "POST":
        query = request.form["search"].lower()  # Get the search query from form input

        wb = openpyxl.load_workbook('Cleaned_Game_Data.xlsx')
        ws = wb["Data_Used_For_Reviews"]

        games_dict = {}  # Dictionary to store games that match the query

        # Assuming each row in Excel represents a game's details
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[25]:
                continue

            game_title = row[2]
            game_details = {
                "trailer_link": row[25],
                "id": row[1],
            }

            if query in game_title.lower():
                games_dict[game_title] = game_details
            print(games_dict)


        # Pass the filtered games dictionary to the template
        return render_template('games.html', results=True,games=games_dict)

    # If it's not a POST request or no query is submitted, render the template normally
    return render_template('games.html')

@app.route('/game/<id>', methods=['GET', 'POST'])
def game(id):
    wb = openpyxl.load_workbook('Cleaned_Game_Data.xlsx')
    ws = wb["Data_Used_For_Reviews"]

    game_data = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        game_title = row[2]
        game_details = {
            "title": game_title,
            "trailer_link": row[25],
            "platform": row[3],
            "release_date": row[4],
            "genre": row[5],
            "publisher": row[6],
            "user_score": row[14],
            "age-rating": row[17],
        }

        if int(id) == row[1]:
            game_data = game_details
            break

    return render_template('game.html', data=game_data)

def sort_songs(data):
    return data[0]

@app.route('/event/<event_id>', methods=['GET', 'POST'])
def event(event_id):
    if request.method == "POST":
        if "leave" in request.form:
            db = sqlite3.connect(db_key)

            cursor = db.cursor()

            cursor.execute('DELETE FROM event_users_'+str(event_id)+' WHERE id = ?', (session["user"]["id"],))
            cursor.execute('DELETE FROM user_events_'+str(session["user"]["id"])+' WHERE id = ?', (event_id,))

            db.commit()
        return redirect(url_for('home'))

    with sqlite3.connect(db_key) as db:
        cursor = db.cursor()

        cursor.execute('SELECT * FROM user_events_'+str(session["user"]["id"])+' WHERE id = ?', (event_id,))
        userEvent = cursor.fetchone()

        if not userEvent:
            return render_template("404.html")

        cursor.execute('SELECT * FROM events WHERE id = ?', (event_id,))
        thisEvent = cursor.fetchone()

        cursor.execute('SELECT * FROM event_users_'+str(thisEvent[0]))
        users = cursor.fetchall()

        cursor.execute('SELECT * FROM event_songs_' + str(thisEvent[0]))
        songs = cursor.fetchall()

        songs.sort(key=sort_songs,reverse=True)

        if not thisEvent:
            return render_template("404.html")

        time = datetime.datetime.fromtimestamp(thisEvent[5])

        return render_template('event.html',event=thisEvent,users=users,songs=songs,time=time)

@app.route('/user/<user_id>')
def user(user_id):
    with sqlite3.connect(db_key) as db:
        cursor = db.cursor()
        cursor.execute('SELECT * FROM users WHERE id = ?', (user_id,))
        thisUser = cursor.fetchone()

        if not thisUser:
            return render_template("404.html")

        return render_template('user.html',user=thisUser)

def generate_code():
    with sqlite3.connect(db_key) as db:
        cursor = db.cursor()

        code = ""

        while code == "":
            letters = string.ascii_lowercase
            code = ''.join(random.choice(letters) for i in range(5))

            cursor.execute('SELECT * FROM events WHERE code = ?', (code.upper(),))
            result = cursor.fetchone()

            if result:
                code = ""

        return code.upper()

@app.route('/create-event', methods=['GET', 'POST'])
def create_event():
    if request.method == "POST":
        datetime_string = request.form["event-time"]
        dt_obj = datetime.datetime.strptime(datetime_string, "%Y-%m-%dT%H:%M")
        unix_timestamp = int(dt_obj.timestamp())

        name = request.form["name"]
        description = request.form["description"]
        owner = session["user"]

        try:
            with sqlite3.connect(db_key) as db:
                cursor = db.cursor()

                query = f"SELECT COUNT(*) FROM events"
                cursor.execute(query)
                result = cursor.fetchone()[0]

                code = generate_code()
                image = "https://media.istockphoto.com/id/519354510/vector/ticket-icon-iconic-series.jpg?s=612x612&w=0&k=20&c=vPPKI5wpK_dNR1FTtOIWqq8uGy7xNv61HLfQvB58Ozk="

                cursor.execute('INSERT INTO events (id, owner_id, owner_name, owner_image, name, unix_time, code, description, image) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', (result + 1, owner["id"], owner["name"], owner["image"], name, unix_timestamp, code, description, image))
                cursor.execute('''CREATE TABLE IF NOT EXISTS event_users_''' + str(result + 1) + ''' (id INTEGER, name TEXT, image TEXT)''')
                cursor.execute('''CREATE TABLE IF NOT EXISTS event_songs_''' + str(result + 1) + ''' (id INTEGER UNIQUE, name TEXT, image TEXT, votes INTEGER)''')
                cursor.execute('''CREATE TABLE IF NOT EXISTS user_''' + str(owner["id"]) + '''event_'''+str(result+1)+''' (id INTEGER)''')

                cursor.execute('INSERT INTO event_users_'+str(result+1)+' (id, name, image) VALUES (?, ?, ?)', (owner["id"], owner["name"], owner["image"]))
                cursor.execute('INSERT INTO user_events_' + str(owner["id"]) + ' (id, name, image) VALUES (?, ?, ?)', (result+1, name, image))

                db.commit()
            return redirect(url_for('home'))
        except sqlite3.IntegrityError:
            flash('error', 'danger')

    return render_template('create_event.html')

@app.route("/join-event", methods=['GET', 'POST'])
def join_event():
    if request.method == "POST":
        code = request.form["code"]

        try:
            with sqlite3.connect(db_key) as db:
                db = sqlite3.connect(db_key)
                cursor = db.cursor()

                cursor.execute('SELECT * FROM events WHERE code = ?', (code,))
                event = cursor.fetchone()

                if not event:
                    flash('Invalid code!', 'danger')
                    return redirect(url_for('join_event'))

                cursor.execute('SELECT * FROM user_events_'+str(session["user"]["id"])+' WHERE id = ?', (event[0],))
                if cursor.fetchone():
                    flash('You are already in this event!', 'danger')
                    return redirect(url_for('join_event'))

                cursor.execute('INSERT INTO event_users_' + str(event[0]) + ' (id, name, image) VALUES (?, ?, ?)',(session["user"]["id"], session["user"]["name"], session["user"]["image"]))
                cursor.execute('INSERT INTO user_events_' + str(session["user"]["id"]) + ' (id, name, image) VALUES (?, ?, ?)',(event[0], event[4], event[8]))
                cursor.execute('''CREATE TABLE IF NOT EXISTS user_''' + str(session["user"]["id"]) + '''event_'''+str(event[0])+''' (id INTEGER)''')

                db.commit()

                flash('Sucessfully joined event!', 'success')
                return redirect(url_for('join_event'))
        except sqlite3.IntegrityError:
            flash('error', 'danger')
            return redirect(url_for('join_event'))

    return render_template('join_event.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        image = "https://static.vecteezy.com/system/resources/previews/008/442/086/original/illustration-of-human-icon-user-symbol-icon-modern-design-on-blank-background-free-vector.jpg"

        try:
            with sqlite3.connect(db_key) as db:
                cursor = db.cursor()

                query = f"SELECT COUNT(*) FROM users"
                cursor.execute(query)
                result = cursor.fetchone()[0]

                cursor.execute('INSERT INTO users (id, username, password, image) VALUES (?, ?, ?, ?)', (result+1, username, hashed_password, image))
                cursor.execute('''CREATE TABLE IF NOT EXISTS user_events_'''+str(result+1)+''' (id INTEGER PRIMARY KEY, name TEXT, image TEXT)''')
                db.commit()
            flash('Registered successfully!', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username already exists!', 'danger')
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with sqlite3.connect(db_key) as db:
            cursor = db.cursor()
            cursor.execute('SELECT * FROM users WHERE username = ?', (username,))
            user = cursor.fetchone()

            if user and check_password_hash(user[2], password):
                session['user'] = {
                    "id": user[0],
                    "name": user[1],
                    "image": user[3],
                    "db_key": db_key,
                }

                return redirect(url_for('home'))
            else:
                flash('Invalid username or password!', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('user_events', None)
    session.pop('selected_event', None)

    return redirect(url_for('home'))

if __name__ == '__main__':
    init_db()

    app.run("0.0.0.0")